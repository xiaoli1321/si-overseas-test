import pytest
from httpx import ASGITransport, AsyncClient

from src.main import app


@pytest.mark.asyncio
async def test_export_records_with_token_query_parameter() -> None:
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        # 1. Login to get a valid token
        login_response = await ac.post(
            "/api/v1/auth/login",
            json={"email": "christest@sibionics.com", "password": "password123"},
        )
        assert login_response.status_code == 200, login_response.text
        token = login_response.json()["data"]["access_token"]

        # 2. Query /api/v1/records/export using Authorization header (should succeed)
        headers = {"Authorization": f"Bearer {token}"}
        resp_hdr = await ac.get("/api/v1/records/export", headers=headers)
        assert resp_hdr.status_code == 200, resp_hdr.text
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp_hdr.headers["content-type"]

        # 3. Query /api/v1/records/export using token query parameter instead of header (should succeed)
        resp_query = await ac.get(f"/api/v1/records/export?token={token}")
        assert resp_query.status_code == 200, resp_query.text
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp_query.headers["content-type"]
        assert "detect-records.xlsx" in resp_query.headers["content-disposition"]

        # 4. Query without token (should fail with 401)
        resp_fail = await ac.get("/api/v1/records/export")
        assert resp_fail.status_code == 401


@pytest.mark.asyncio
async def test_export_records_excel_content() -> None:
    import openpyxl
    from io import BytesIO
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        # 1. Login to get a valid token
        login_response = await ac.post(
            "/api/v1/auth/login",
            json={"email": "christest@sibionics.com", "password": "password123"},
        )
        assert login_response.status_code == 200
        token = login_response.json()["data"]["access_token"]

        # 2. Export records
        headers = {"Authorization": f"Bearer {token}"}
        response = await ac.get("/api/v1/records/export", headers=headers)
        assert response.status_code == 200

        # 3. Parse Excel content
        content = response.content
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]

        assert len(rows) > 0
        header_row = rows[0]
        expected_headers = [
            "Device Identifier",
            "Fault Category",
            "Fault Subtype",
            "Status",
            "Verdict",
            "Feedback",
            "What we found",
            "Judging after-sales standards",
            "Possible causes",
            "Created At",
        ]
        assert header_row == expected_headers

        # Check that rows have the correct number of fields
        for row in rows[1:]:
            assert len(row) == len(expected_headers)

        # Verify that column widths are correctly defined (non-default)
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(expected_headers) + 1):
            col_letter = get_column_letter(col_idx)
            col_dim = ws.column_dimensions.get(col_letter)
            assert col_dim is not None, f"Column dimension for {col_letter} should be set"
            assert col_dim.width is not None, f"Width for column {col_letter} should be set"
            assert 12 <= col_dim.width <= 50, f"Column {col_letter} width {col_dim.width} should be within [12, 50]"

        # 4. Feedback a record as rejected with a reason
        list_resp = await ac.get("/api/v1/records?page=1&page_size=5", headers=headers)
        assert list_resp.status_code == 200
        records_list = list_resp.json()["data"]["items"]
        assert len(records_list) > 0
        record_id = records_list[0]["id"]

        test_reason = "Test rejection reason explanation"
        feedback_resp = await ac.post(
            f"/api/v1/records/{record_id}/feedback",
            json={"feedback_status": "rejected", "reject_reason": test_reason},
            headers=headers,
        )
        assert feedback_resp.status_code == 200

        # 5. Re-export and verify rejected reason format in Excel content
        response2 = await ac.get("/api/v1/records/export", headers=headers)
        assert response2.status_code == 200
        wb2 = openpyxl.load_workbook(BytesIO(response2.content))
        ws2 = wb2.active
        rows2 = [[cell.value for cell in row] for row in ws2.iter_rows()]

        # Find the row for the updated record and assert its feedback column matches
        found_updated = False
        for row in rows2[1:]:
            if row[5] and f"rejected: {test_reason}" in str(row[5]):
                found_updated = True
                break
        assert found_updated, "Rejection reason was not found in the Feedback column of the Excel export"


