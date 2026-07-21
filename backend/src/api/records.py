from datetime import datetime
from io import BytesIO
import logging
from typing import Annotated

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from src.api.deps import get_current_user
from src.core.database import get_db
from src.core.exceptions import NotFoundError
from src.core.logging import log_context
from src.core.responses import ok
from src.models.tables import User
from src.repositories.store import (
    get_record,
    get_user_by_id,
    get_users_by_ids,
    iter_records_for_export,
    list_records,
    stats,
)
from src.schemas.common import Page
from src.schemas.domain import (
    BatchDeleteRequest,
    DashboardStatsResponse,
    DetectRecordResponse,
    FeedbackRequest,
)
from src.schemas.frontend import (
    record_to_frontend,
    record_to_list_item,
    stats_to_frontend,
)
from src.services.audit import record_audit_event
from src.services.analytics import track_verdict_adoption
from src.services.detections import batch_delete_records, delete_record, update_feedback

router = APIRouter(prefix="/records", tags=["records"])
logger = logging.getLogger(__name__)


def _parse_date_from(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def _parse_date_to(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(value)
        return dt.replace(hour=23, minute=59, second=59, microsecond=999999)
    except ValueError:
        return None


def _display_width(value: str) -> int:
    return sum(2 if ord(char) > 127 else 1 for char in value)


@router.get("")
async def list_endpoint(
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    fault_category: str | None = None,
    verdict: str | None = None,
    conclusion: str | None = None,
    serial_no: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    account_id: int | None = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
) -> dict:
    records, total = await list_records(
        db,
        user,
        source="web",
        fault_category=fault_category,
        verdict=verdict,
        serial_no=serial_no,
        date_from=_parse_date_from(date_from),
        date_to=_parse_date_to(date_to),
        conclusion=conclusion,
        account_id=account_id,
        page=page,
        page_size=page_size,
    )
    submitters = await get_users_by_ids(db, [record.user_id for record in records])
    items = [
        record_to_list_item(record, submitters.get(record.user_id))
        for record in records
    ]
    logger.info(
        "Detection records listed",
        extra=log_context(
            "records.listed",
            user_id=user.id,
            result_count=len(records),
            total=total,
            page=page,
            page_size=page_size,
        ),
    )
    return ok(
        Page(items=items, total=total, page=page, page_size=page_size).model_dump()
    )


@router.get("/stats")
async def stats_endpoint(
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    data = stats_to_frontend(await stats(db, user, source="web"))
    logger.info(
        "Detection record stats loaded",
        extra=log_context("records.stats_loaded", user_id=user.id),
    )
    return ok(data)


@router.get("/export")
async def export_endpoint(
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    fault_category: str | None = None,
    verdict: str | None = None,
    conclusion: str | None = None,
    serial_no: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    account_id: int | None = None,
) -> StreamingResponse:
    import openpyxl
    from openpyxl.utils import get_column_letter

    headers = [
        "Device Identifier",
        "Fault Category",
        "Fault Subtype",
        "Status",
        "Verdict",
        "Feedback",
        "What we found",
        "Judging after-sales standards",
        "Possible causes",
        "Created At",
    ]
    rows = []
    column_widths = [len(header) for header in headers]
    exported_count = 0

    from src.rules.presentation import build_verdict_presentation
    from src.schemas.frontend import frontend_threshold_profile

    async for records in iter_records_for_export(
        db,
        user,
        source="web",
        fault_category=fault_category,
        verdict=verdict,
        serial_no=serial_no,
        date_from=_parse_date_from(date_from),
        date_to=_parse_date_to(date_to),
        conclusion=conclusion,
        account_id=account_id,
    ):
        for record in records:
            profile = record.threshold_snapshot or {
                "version": 1,
                "savedAt": None,
                "rules": {},
            }
            threshold_profile = (
                profile
                if "rules" in profile
                else frontend_threshold_profile(profile, record.threshold_id or 1, None)
            )

            presentation = build_verdict_presentation(
                fault_category=record.fault_category,
                fault_subtype=record.fault_subtype,
                verdict=record.verdict,
                issue_detected=record.issue_detected,
                evidence=record.evidence,
                threshold_snapshot=threshold_profile,
            )
            title = presentation.get("title") if presentation else None
            if title and title != "/":
                display_subtype = title
            else:
                display_subtype = record.fault_subtype or ""

            what_we_found = presentation.get("whatWeFound") if presentation else ""
            if not what_we_found or what_we_found == "/":
                what_we_found = ""

            why_this_result = presentation.get("whyThisResult") if presentation else ""
            if not why_this_result or why_this_result == "/":
                why_this_result = ""

            possible_causes = presentation.get("possibleCauses") if presentation else ""
            if not possible_causes or possible_causes == "/":
                possible_causes = ""

            feedback_val = record.feedback_status
            if feedback_val == "rejected" and record.reject_reason:
                feedback_val = f"rejected: {record.reject_reason}"

            row = [
                record.serial_no,
                record.fault_category,
                display_subtype,
                record.status,
                record.verdict,
                feedback_val,
                what_we_found,
                why_this_result,
                possible_causes,
                record.created_at.strftime("%Y-%m-%d %H:%M:%S")
                if record.created_at
                else "",
            ]
            rows.append(row)
            exported_count += 1
            for index, value in enumerate(row):
                value_width = max(
                    _display_width(line) for line in str(value or "").split("\n")
                )
                column_widths[index] = max(column_widths[index], value_width)

    logger.info(
        "Detection records exported",
        extra=log_context(
            "records.exported", user_id=user.id, record_count=exported_count
        ),
    )

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Detect Records")
    ws.title = "Detect Records"

    for index, max_len in enumerate(column_widths, start=1):
        col_letter = get_column_letter(index)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    ws.append(headers)
    for row in rows:
        ws.append(row)

    out_buf = BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)

    return StreamingResponse(
        out_buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=detect-records.xlsx"},
    )


@router.get("/{record_id}")
async def detail_endpoint(
    record_id: int,
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    record = await get_record(db, user.id, record_id, source="web", viewer=user)
    if record is None:
        raise NotFoundError("Detect record was not found.")
    submitter = await get_user_by_id(db, record.user_id)
    logger.info(
        "Detection record loaded",
        extra=log_context("records.loaded", user_id=user.id, record_id=record_id),
    )
    return ok(record_to_frontend(record, submitter or user))


@router.post("/{record_id}/feedback")
async def feedback_endpoint(
    record_id: int,
    payload: FeedbackRequest,
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    feedback_status = payload.feedback_status
    if feedback_status is None:
        feedback_status = "adopted" if payload.verdict_adoption == "Yes" else "rejected"
    record = await update_feedback(
        db,
        user_id=user.id,
        record_id=record_id,
        feedback_status=feedback_status,
        reject_reason=payload.reject_reason,
    )
    await record_audit_event(
        db,
        user_id=user.id,
        action="detection.feedback",
        target_type="detect_record",
        target_id=record.id,
        metadata={
            "feedback_status": record.feedback_status,
            "reject_reason": record.reject_reason,
            "adopted_at": record.adopted_at.isoformat() if record.adopted_at else None,
        },
    )
    await track_verdict_adoption(
        db,
        user=user,
        record=record,
        feedback_status=feedback_status,
        reject_reason=payload.reject_reason,
    )
    await db.commit()
    logger.info(
        "Detection record feedback updated",
        extra=log_context(
            "records.feedback_updated",
            user_id=user.id,
            record_id=record.id,
            feedback_status=feedback_status,
        ),
    )
    return ok(record_to_frontend(record, user))


@router.delete("/{record_id}")
async def delete_endpoint(
    record_id: int,
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    record = await delete_record(db, user_id=user.id, record_id=record_id)
    await record_audit_event(
        db,
        user_id=user.id,
        action="detection.delete",
        target_type="detect_record",
        target_id=record.id,
        metadata={
            "serial_no": record.serial_no,
            "fault_category": record.fault_category,
        },
    )
    await db.commit()
    logger.info(
        "Detection record deleted",
        extra=log_context(
            "records.deleted",
            user_id=user.id,
            record_id=record.id,
        ),
    )
    return ok(
        {"id": str(record.id), "isVisibleInWorkbench": record.is_visible_in_workbench}
    )


@router.post("/batch-delete")
async def batch_delete_endpoint(
    payload: BatchDeleteRequest,
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    deleted_ids = await batch_delete_records(
        db, user_id=user.id, record_ids=payload.record_ids
    )
    if deleted_ids:
        await record_audit_event(
            db,
            user_id=user.id,
            action="detection.batch_delete",
            target_type="detect_record",
            status="success",
            metadata={
                "count": len(deleted_ids),
                "deleted_ids": deleted_ids,
            },
        )
        await db.commit()
    logger.info(
        "Detection records batch-deleted",
        extra=log_context(
            "records.batch_deleted",
            user_id=user.id,
            requested_count=len(payload.record_ids),
            deleted_count=len(deleted_ids),
        ),
    )
    return ok({"deletedIds": [str(x) for x in deleted_ids]})
